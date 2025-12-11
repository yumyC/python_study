"""
工作日志相关的异步任务

提供工作日志导出等异步处理功能
"""

import os
import tempfile
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from .celery_app import celery
from app.models import WorkLog, Employee


@celery.task(bind=True)
def export_work_logs_task(self, employee_id=None, start_date=None, end_date=None, user_id=None):
    """
    导出工作日志到 Excel 文件
    
    Args:
        employee_id: 员工ID（可选）
        start_date: 开始日期（可选）
        end_date: 结束日期（可选）
        user_id: 请求用户ID
        
    Returns:
        dict: 包含文件路径和相关信息的字典
    """
    try:
        # 更新任务状态
        self.update_state(
            state='PROGRESS',
            meta={'progress': 10, 'message': '开始查询数据...'}
        )
        
        # 构建查询
        query = WorkLog.query
        
        # 员工筛选
        if employee_id:
            query = query.filter(WorkLog.employee_id == employee_id)
        
        # 日期筛选
        if start_date:
            try:
                start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
                query = query.filter(WorkLog.log_date >= start_date)
            except ValueError:
                pass
        
        if end_date:
            try:
                end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
                query = query.filter(WorkLog.log_date <= end_date)
            except ValueError:
                pass
        
        # 获取数据
        work_logs = query.order_by(WorkLog.log_date.desc()).all()
        
        # 更新任务状态
        self.update_state(
            state='PROGRESS',
            meta={'progress': 30, 'message': f'查询到 {len(work_logs)} 条记录，开始生成Excel...'}
        )
        
        # 创建 Excel 工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "工作日志"
        
        # 设置表头
        headers = [
            "员工姓名", "日期", "工作内容", "完成状态", "遇到的问题", 
            "明日计划", "自我评分", "上级评分", "上级评价", "创建时间"
        ]
        
        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 更新任务状态
        self.update_state(
            state='PROGRESS',
            meta={'progress': 50, 'message': '正在写入数据...'}
        )
        
        # 写入数据
        for row, log in enumerate(work_logs, 2):
            export_data = log.to_export_dict()
            
            for col, header in enumerate(headers, 1):
                value = export_data.get(header, "")
                ws.cell(row=row, column=col, value=value)
            
            # 更新进度
            if row % 100 == 0:
                progress = 50 + (row / len(work_logs)) * 40
                self.update_state(
                    state='PROGRESS',
                    meta={'progress': int(progress), 'message': f'已处理 {row-1} 条记录...'}
                )
        
        # 调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 更新任务状态
        self.update_state(
            state='PROGRESS',
            meta={'progress': 90, 'message': '正在保存文件...'}
        )
        
        # 生成文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"work_logs_export_{timestamp}.xlsx"
        
        # 创建临时文件
        temp_dir = tempfile.gettempdir()
        file_path = os.path.join(temp_dir, filename)
        
        # 保存文件
        wb.save(file_path)
        
        # 返回结果
        result = {
            'file_path': file_path,
            'filename': filename,
            'record_count': len(work_logs),
            'export_time': datetime.now().isoformat(),
            'user_id': user_id
        }
        
        return result
        
    except Exception as e:
        # 任务失败
        self.update_state(
            state='FAILURE',
            meta={'error': str(e), 'message': '导出失败'}
        )
        raise e


@celery.task
def cleanup_export_files():
    """
    清理过期的导出文件
    
    定期清理临时目录中的导出文件
    """
    try:
        import glob
        from datetime import timedelta
        
        temp_dir = tempfile.gettempdir()
        pattern = os.path.join(temp_dir, "work_logs_export_*.xlsx")
        
        # 获取所有导出文件
        files = glob.glob(pattern)
        
        # 删除超过24小时的文件
        cutoff_time = datetime.now() - timedelta(hours=24)
        deleted_count = 0
        
        for file_path in files:
            try:
                # 获取文件修改时间
                file_time = datetime.fromtimestamp(os.path.getmtime(file_path))
                
                if file_time < cutoff_time:
                    os.remove(file_path)
                    deleted_count += 1
                    
            except Exception as e:
                print(f"Failed to delete file {file_path}: {str(e)}")
        
        return {
            'deleted_count': deleted_count,
            'cleanup_time': datetime.now().isoformat()
        }
        
    except Exception as e:
        print(f"Cleanup task failed: {str(e)}")
        raise e


# 配置定期任务
from celery.schedules import crontab

celery.conf.beat_schedule = {
    'cleanup-export-files': {
        'task': 'app.tasks.work_log_tasks.cleanup_export_files',
        'schedule': crontab(hour=2, minute=0),  # 每天凌晨2点执行
    },
}