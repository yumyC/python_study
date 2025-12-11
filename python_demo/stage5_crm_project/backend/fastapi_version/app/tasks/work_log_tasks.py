"""
工作日志相关异步任务

实现工作日志的异步导出功能，支持 Excel 格式导出
"""

from celery import current_task
from sqlalchemy.orm import Session
from sqlalchemy import and_, or_
from datetime import datetime, date
from typing import Optional, List, Dict, Any
import os
import tempfile
import uuid
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .celery_app import celery_app
from ..database import SessionLocal
from ..models import WorkLog, Employee, Position


@celery_app.task(bind=True, name="export_work_logs_task")
def export_work_logs_task(
    self,
    employee_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    export_format: str = "xlsx",
    requested_by: Optional[str] = None
) -> Dict[str, Any]:
    """
    异步导出工作日志任务
    
    Args:
        self: Celery 任务实例
        employee_id: 员工ID（可选，为空则导出所有员工）
        start_date: 开始日期（YYYY-MM-DD 格式）
        end_date: 结束日期（YYYY-MM-DD 格式）
        export_format: 导出格式（目前支持 xlsx）
        requested_by: 请求导出的用户ID
        
    Returns:
        导出结果信息
    """
    try:
        # 更新任务状态
        current_task.update_state(
            state="PROGRESS",
            meta={"current": 0, "total": 100, "status": "开始导出工作日志..."}
        )
        
        # 创建数据库会话
        db = SessionLocal()
        
        try:
            # 构建查询条件
            query = db.query(WorkLog).join(Employee)
            
            # 按员工筛选
            if employee_id:
                query = query.filter(WorkLog.employee_id == employee_id)
            
            # 按日期范围筛选
            if start_date:
                start_dt = datetime.strptime(start_date, "%Y-%m-%d").date()
                query = query.filter(WorkLog.log_date >= start_dt)
            
            if end_date:
                end_dt = datetime.strptime(end_date, "%Y-%m-%d").date()
                query = query.filter(WorkLog.log_date <= end_dt)
            
            # 按日期排序
            work_logs = query.order_by(WorkLog.log_date.desc(), Employee.full_name).all()
            
            current_task.update_state(
                state="PROGRESS",
                meta={"current": 20, "total": 100, "status": f"查询到 {len(work_logs)} 条工作日志"}
            )
            
            if not work_logs:
                return {
                    "success": False,
                    "message": "没有找到符合条件的工作日志",
                    "file_path": None
                }
            
            # 生成 Excel 文件
            file_path = _generate_excel_file(work_logs, current_task)
            
            current_task.update_state(
                state="PROGRESS",
                meta={"current": 90, "total": 100, "status": "生成下载链接..."}
            )
            
            # 生成文件信息
            file_size = os.path.getsize(file_path)
            file_name = os.path.basename(file_path)
            
            result = {
                "success": True,
                "message": f"成功导出 {len(work_logs)} 条工作日志",
                "file_path": file_path,
                "file_name": file_name,
                "file_size": file_size,
                "record_count": len(work_logs),
                "export_time": datetime.now().isoformat(),
                "requested_by": requested_by,
                "filters": {
                    "employee_id": employee_id,
                    "start_date": start_date,
                    "end_date": end_date
                }
            }
            
            current_task.update_state(
                state="SUCCESS",
                meta={"current": 100, "total": 100, "status": "导出完成"}
            )
            
            return result
            
        finally:
            db.close()
            
    except Exception as exc:
        # 记录错误并返回失败状态
        error_msg = f"导出工作日志失败: {str(exc)}"
        
        current_task.update_state(
            state="FAILURE",
            meta={"error": error_msg, "traceback": str(exc)}
        )
        
        return {
            "success": False,
            "message": error_msg,
            "file_path": None
        }


def _generate_excel_file(work_logs: List[WorkLog], task=None) -> str:
    """
    生成 Excel 文件
    
    Args:
        work_logs: 工作日志列表
        task: Celery 任务实例（用于更新进度）
        
    Returns:
        生成的文件路径
    """
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "工作日志"
    
    # 设置列标题
    headers = [
        "序号", "员工姓名", "岗位", "日志日期", "工作内容", 
        "完成状态", "遇到问题", "明日计划", "自评分数", 
        "上级评分", "上级评语", "创建时间"
    ]
    
    # 写入标题行
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
    
    # 写入数据行
    total_logs = len(work_logs)
    for idx, log in enumerate(work_logs, 1):
        row = idx + 1
        
        # 更新进度
        if task and idx % 50 == 0:  # 每处理50条记录更新一次进度
            progress = 20 + int((idx / total_logs) * 60)  # 20-80% 的进度用于数据处理
            task.update_state(
                state="PROGRESS",
                meta={"current": progress, "total": 100, "status": f"处理数据 {idx}/{total_logs}"}
            )
        
        # 填充数据
        data = [
            idx,  # 序号
            log.employee.full_name if log.employee else "未知",
            log.employee.position.name if log.employee and log.employee.position else "未设置",
            log.log_date.strftime("%Y-%m-%d") if log.log_date else "",
            log.work_content or "",
            log.get_status_display(),
            log.problems_encountered or "",
            log.tomorrow_plan or "",
            log.self_rating or "",
            log.supervisor_rating or "",
            log.supervisor_comment or "",
            log.created_at.strftime("%Y-%m-%d %H:%M:%S") if log.created_at else ""
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
            
            # 根据完成状态设置颜色
            if col == 6:  # 完成状态列
                if value == "已完成":
                    cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                elif value == "进行中":
                    cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                elif value == "待开始":
                    cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    
    # 调整列宽
    column_widths = [8, 15, 15, 12, 30, 12, 25, 25, 10, 10, 20, 18]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # 冻结首行
    ws.freeze_panes = "A2"
    
    # 生成文件名和路径
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"work_logs_export_{timestamp}_{uuid.uuid4().hex[:8]}.xlsx"
    
    # 创建临时目录（如果不存在）
    export_dir = os.path.join(tempfile.gettempdir(), "crm_exports")
    os.makedirs(export_dir, exist_ok=True)
    
    file_path = os.path.join(export_dir, file_name)
    
    # 保存文件
    wb.save(file_path)
    
    if task:
        task.update_state(
            state="PROGRESS",
            meta={"current": 85, "total": 100, "status": "Excel 文件生成完成"}
        )
    
    return file_path


@celery_app.task(name="cleanup_export_files")
def cleanup_export_files_task(max_age_hours: int = 24) -> Dict[str, Any]:
    """
    清理过期的导出文件
    
    Args:
        max_age_hours: 文件最大保留时间（小时）
        
    Returns:
        清理结果
    """
    try:
        export_dir = os.path.join(tempfile.gettempdir(), "crm_exports")
        
        if not os.path.exists(export_dir):
            return {"success": True, "message": "导出目录不存在", "deleted_count": 0}
        
        now = datetime.now()
        deleted_count = 0
        total_size = 0
        
        for filename in os.listdir(export_dir):
            file_path = os.path.join(export_dir, filename)
            
            if os.path.isfile(file_path):
                # 检查文件年龄
                file_mtime = datetime.fromtimestamp(os.path.getmtime(file_path))
                age_hours = (now - file_mtime).total_seconds() / 3600
                
                if age_hours > max_age_hours:
                    file_size = os.path.getsize(file_path)
                    os.remove(file_path)
                    deleted_count += 1
                    total_size += file_size
        
        return {
            "success": True,
            "message": f"清理完成，删除了 {deleted_count} 个过期文件",
            "deleted_count": deleted_count,
            "freed_space": total_size,
            "max_age_hours": max_age_hours
        }
        
    except Exception as exc:
        return {
            "success": False,
            "message": f"清理文件失败: {str(exc)}",
            "deleted_count": 0
        }


@celery_app.task(name="generate_work_log_report")
def generate_work_log_report_task(
    start_date: str,
    end_date: str,
    department_id: Optional[str] = None
) -> Dict[str, Any]:
    """
    生成工作日志统计报告
    
    Args:
        start_date: 开始日期
        end_date: 结束日期
        department_id: 部门ID（可选）
        
    Returns:
        报告生成结果
    """
    try:
        db = SessionLocal()
        
        try:
            # 解析日期
            start_dt = datetime.strptime(start_date, "%Y-%m-%d").date()
            end_dt = datetime.strptime(end_date, "%Y-%m-%d").date()
            
            # 构建查询
            query = db.query(WorkLog).join(Employee)
            query = query.filter(
                and_(
                    WorkLog.log_date >= start_dt,
                    WorkLog.log_date <= end_dt
                )
            )
            
            if department_id:
                query = query.join(Position).filter(Position.id == department_id)
            
            work_logs = query.all()
            
            # 生成统计数据
            stats = _calculate_work_log_statistics(work_logs)
            
            return {
                "success": True,
                "period": {"start_date": start_date, "end_date": end_date},
                "statistics": stats,
                "generated_at": datetime.now().isoformat()
            }
            
        finally:
            db.close()
            
    except Exception as exc:
        return {
            "success": False,
            "message": f"生成报告失败: {str(exc)}"
        }


def _calculate_work_log_statistics(work_logs: List[WorkLog]) -> Dict[str, Any]:
    """
    计算工作日志统计数据
    
    Args:
        work_logs: 工作日志列表
        
    Returns:
        统计数据
    """
    if not work_logs:
        return {
            "total_logs": 0,
            "completion_stats": {},
            "rating_stats": {},
            "employee_stats": {}
        }
    
    # 基础统计
    total_logs = len(work_logs)
    
    # 完成状态统计
    completion_stats = {}
    for log in work_logs:
        status = log.completion_status.value
        completion_stats[status] = completion_stats.get(status, 0) + 1
    
    # 评分统计
    self_ratings = [log.self_rating for log in work_logs if log.self_rating]
    supervisor_ratings = [log.supervisor_rating for log in work_logs if log.supervisor_rating]
    
    rating_stats = {
        "self_rating": {
            "count": len(self_ratings),
            "average": sum(self_ratings) / len(self_ratings) if self_ratings else 0,
            "min": min(self_ratings) if self_ratings else 0,
            "max": max(self_ratings) if self_ratings else 0
        },
        "supervisor_rating": {
            "count": len(supervisor_ratings),
            "average": sum(supervisor_ratings) / len(supervisor_ratings) if supervisor_ratings else 0,
            "min": min(supervisor_ratings) if supervisor_ratings else 0,
            "max": max(supervisor_ratings) if supervisor_ratings else 0
        }
    }
    
    # 员工统计
    employee_stats = {}
    for log in work_logs:
        emp_name = log.employee.full_name if log.employee else "未知"
        if emp_name not in employee_stats:
            employee_stats[emp_name] = {
                "log_count": 0,
                "avg_self_rating": 0,
                "avg_supervisor_rating": 0,
                "completion_rate": 0
            }
        
        employee_stats[emp_name]["log_count"] += 1
    
    # 计算每个员工的详细统计
    for emp_name in employee_stats:
        emp_logs = [log for log in work_logs if log.employee and log.employee.full_name == emp_name]
        
        # 平均评分
        emp_self_ratings = [log.self_rating for log in emp_logs if log.self_rating]
        emp_supervisor_ratings = [log.supervisor_rating for log in emp_logs if log.supervisor_rating]
        
        employee_stats[emp_name]["avg_self_rating"] = (
            sum(emp_self_ratings) / len(emp_self_ratings) if emp_self_ratings else 0
        )
        employee_stats[emp_name]["avg_supervisor_rating"] = (
            sum(emp_supervisor_ratings) / len(emp_supervisor_ratings) if emp_supervisor_ratings else 0
        )
        
        # 完成率
        completed_logs = [log for log in emp_logs if log.completion_status.value == "completed"]
        employee_stats[emp_name]["completion_rate"] = (
            len(completed_logs) / len(emp_logs) * 100 if emp_logs else 0
        )
    
    return {
        "total_logs": total_logs,
        "completion_stats": completion_stats,
        "rating_stats": rating_stats,
        "employee_stats": employee_stats
    }